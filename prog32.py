import urllib.request
import urllib.parse
from html.parser import HTMLParser
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

class SinoptikParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.min_temps = []
        self.max_temps = []
        self.in_min = False
        self.in_max = False
        self.capture_data = False

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == 'div' and 'class' in attrs_dict:
            if attrs_dict['class'] == 'min':
                self.in_min = True
            elif attrs_dict['class'] == 'max':
                self.in_max = True
        if tag == 'span' and (self.in_min or self.in_max):
            self.capture_data = True

    def handle_endtag(self, tag):
        if tag == 'span':
            self.capture_data = False
        if tag == 'div':
            self.in_min = False
            self.in_max = False

    def handle_data(self, data):
        if self.capture_data:
            clean_data = data.strip()
            if clean_data:
                if self.in_min:
                    self.min_temps.append(clean_data)
                elif self.in_max:
                    self.max_temps.append(clean_data)

def get_weather_html_parser(city):
    city_encoded = urllib.parse.quote(city.lower())
    url = f"https://sinoptik.ua/погода-{city_encoded}"
    
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    req = urllib.request.Request(url, headers=headers)
    
    try:
        with urllib.request.urlopen(req) as response:
            html = response.read().decode('utf-8')
    except Exception as e:
        print(f"Помилка завантаження сторінки: {e}")
        return None

    parser = SinoptikParser()
    parser.feed(html)

    forecast = []
    for i in range(1, 6):
        if i < len(parser.min_temps) and i < len(parser.max_temps):
            forecast.append({
                'day': i,
                'min': parser.min_temps[i],
                'max': parser.max_temps[i]
            })
    return forecast

def save_to_excel(city, data):
    filename = "weather_parser_forecast.xlsx"
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["Дата запиту", "Місто"]
        for i in range(1, 6):
            headers.extend([f"День {i} Мін", f"День {i} Макс"])
        ws.append(headers)
        
    row = [current_date, city.capitalize()]
    for day_data in data:
        row.extend([day_data['min'], day_data['max']])
        
    ws.append(row)
    wb.save(filename)
    print(f"Дані успішно збережено у файл {filename}")

if __name__ == "__main__":
    city_name = "киев"
    print(f"Отримуємо прогноз (HTMLParser) для міста: {city_name}...")
    weather_data = get_weather_html_parser(city_name)
    
    if weather_data and len(weather_data) == 5:
        save_to_excel(city_name, weather_data)
    else:
        print("Не вдалося зібрати повні дані на 5 днів за допомогою структурованого аналізу.")