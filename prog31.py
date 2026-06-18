import urllib.request
import re
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

def get_weather_regex(city):
    city_encoded = urllib.parse.quote(city.lower())
    url = f"https://sinoptik.ua/погода-{city_encoded}"
    
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
    req = urllib.request.Request(url, headers=headers)
    
    try:
        with urllib.request.urlopen(req) as response:
            html = response.read().decode('utf-8')
    except Exception as e:
        print(f"Помилка завантаження сторінки: {e}")
        return None

    min_temps = re.findall(r'<div class="min"><span>([^<]+)</span>', html)
    max_temps = re.findall(r'<div class="max"><span>([^<]+)</span>', html)
    
    forecast = []
    for i in range(1, 6):
        if i < len(min_temps) and i < len(max_temps):
            forecast.append({
                'day': i,
                'min': min_temps[i].strip(),
                'max': max_temps[i].strip()
            })
            
    return forecast

def save_to_excel(city, data):
    filename = "weather_forecast.xlsx"
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["Дата запиту", "Місто"]
        for i in range(1, 6):
            headers.extend([f"День {i} Мін", f"День {i} Макс"])
        ws.append(headers)

    row = [current_date, city.capitalize()]
    for day_data in data:
        row.extend([day_data['min'], day_data['max']])
        
    ws.append(row)
    wb.save(filename)
    print(f"Дані успішно збережено у файл {filename}")

if __name__ == "__main__":
    city_name = "киев"  # можна ввести "львов", "одесса" тощо
    print(f"Отримуємо прогноз (Regex) для міста: {city_name}...")
    weather_data = get_weather_regex(city_name)
    
    if weather_data and len(weather_data) == 5:
        save_to_excel(city_name, weather_data)
    else:
        print("Не вдалося зібрати повні дані на 5 днів.")